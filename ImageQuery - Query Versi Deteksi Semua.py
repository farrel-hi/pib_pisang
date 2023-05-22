import numpy as np
import h5py
import gc
import matplotlib.pyplot as plt
import matplotlib.image as mpimg

import openpyxl
from openpyxl import Workbook




from keras.preprocessing import image

import tensorflow as tf

from tensorflow.keras.models import Model

import os


print(tf.__version__)
print(tf.config.list_physical_devices())

gpus = tf.config.experimental.list_physical_devices('GPU')
#tf.config.experimental.set_virtual_device_configuration(
#         gpus[0],
#            [tf.config.experimental.VirtualDeviceConfiguration(memory_limit=3300)])

config = tf.config.experimental.set_memory_growth(gpus[0], True)

os.environ['CUDA_VISIBLE_DEVICES'] = '0'
def ekstrak(img, cnn):
    if cnn == "ResNet50V2":
        from keras.applications.resnet_v2 import preprocess_input
        from keras.applications.resnet_v2 import ResNet50V2


        img = image.load_img(img, target_size=(224, 224), color_mode='rgb')
        x = image.img_to_array(img)
        x = np.expand_dims(x, axis=0)
        x = preprocess_input(x)


        base_model = ResNet50V2(weights='imagenet')

    elif cnn == "MobileNetV2":
        from keras.applications.mobilenet_v2 import preprocess_input
        from keras.applications.mobilenet_v2 import MobileNetV2

        img = image.load_img(img, target_size=(224, 224), color_mode='rgb')
        x = image.img_to_array(img)
        x = np.expand_dims(x, axis=0)
        x = preprocess_input(x)
        base_model = MobileNetV2(weights='imagenet')

    elif cnn == "EfficientNetB7":
        from tensorflow.keras.applications.efficientnet import preprocess_input
        from tensorflow.keras.applications.efficientnet import EfficientNetB7

        img = image.load_img(img, target_size=(600, 600), color_mode='rgb')

        x = image.img_to_array(img)
        x = np.expand_dims(x, axis=0)
        x = preprocess_input(x)
        base_model = EfficientNetB7(weights='imagenet')

    elif cnn == "InceptionV3":
        from keras.applications.inception_v3 import preprocess_input
        from keras.applications.inception_v3 import InceptionV3

        img = image.load_img(img, target_size=(299, 299), color_mode='rgb')

        x = image.img_to_array(img)
        x = np.expand_dims(x, axis=0)
        x = preprocess_input(x)

        base_model = InceptionV3(weights='imagenet')


    #base_model.summary()
    model = Model(inputs=base_model.input, outputs=base_model.layers[-2].output) #NASNET
    #model = Model(inputs=base_model.input, outputs=base_model.get_layer('global_average_pooling2d_1').output)  #???
    feature = model(x)[0]
    norm_feat = feature / np.linalg.norm(feature)
    return norm_feat



#=============SETTINGS=============

perubahan = 25 #1 Gambar Raw Menghasil X Gambar Preproses =================================================================================
cap = 70 #Persentase minimal untuk menjadi output ===================================================
enable_lock = 0 #Tulis "1" jika merasa output kurang meuaskan (Harus liat berapa yang dibutuhkan)
lock = 0 #Iterasi awal #JANGAN DIUBAHH!!!!!!!
stop = 1 #Lock diiterasi berapa (Harus liat berapa yang dibutuhkan
preprocess = "Prewitt"
cnn = "MobileNetV2" #============================================================================================
pil = 1

#=============END=============


path3 = 'Hasil Preproses - Kulit - '+preprocess+' - Test'+'\\'
h5f = h5py.File("Model_TemuKembali\\TemuKembaliKulit5"+preprocess+cnn+".h5", 'r')
hit2 = 0 #database yang didapat
hit = 0  # iterasi jumlah yang didapat
topcap = cap+1

#KASUS MINKOWSKI 56 CAP = 46.42857142857143

# feats = h5f['dataset_1'][:]
feats = h5f['dataset_1'][:]
#print(feats)
print(feats.shape)
print(feats.dtype)
imgNames = h5f['dataset_2'][:]
#print(imgNames)
#print(imgNames.shape)
h5f.close()


import os
list_gambar = ""

#path2 = root.directory

path2 = 'gambarmentah - balanced\\'
l = os.listdir(path2)

print(path2)

with open("ListGambar.txt", "w", encoding="utf-8") as file:
    for eachfile in l:
        if eachfile.endswith(('.png','.jpg','.jpeg')) == False:
            continue
        list_gambar += eachfile + "\n"
    file.write(list_gambar)

gambar = []
with open('ListGambar.txt', newline='') as infile:
    for line in infile:
        gambar.extend(line.strip().split(','))

#alamatinput = root.directory + '\\'

alamatinput = path = path3

datab = {}
#35
urutan=-1
maxres = perubahan
# 56 #64 #112 #224 #448
# 8  #8  #8   #16  #28
# 8  #9  #15  #15  #18
if perubahan==56:
    columns = 8
    rows = 8
elif perubahan==64:
    columns = 8
    rows = 9
elif perubahan==112:
    columns = 8
    rows = 15
elif perubahan==224:
    columns = 16
    rows = 15
elif perubahan==448:
    columns = 28
    rows = 18


elif perubahan==25:
    columns = 7
    rows = 5
elif perubahan==50:
    columns = 9
    rows = 18
elif perubahan==125:
    columns = 5
    rows = 26


columnsz = int(columns / 2) #MEMANG COLUMNSZ

if cnn == "ResNet50V2":
    modelkeras = 1
elif cnn == "MobileNetV2":
    modelkeras = 2
elif cnn == "EfficientNetB7":
    modelkeras = 3
elif cnn == "InceptionV3":
    modelkeras = 4


if pil==1:
    mode = 'Manhattan'
elif pil==2:
    mode = 'Euclidean'
elif pil==3:
    mode = 'Chebyshev'
elif pil==4:
    mode = 'Minkowski'
elif pil==5:
    mode = 'Vector Cosine'



outquery = 'Data Query - '+preprocess+'\\Query - '+str(maxres)+' '+mode+' '+ cnn+'\\'

if not os.path.exists(outquery):
    os.makedirs(outquery+'Output Text Query'+'\\')
    os.makedirs(outquery + 'Output Akurasi Query' + '\\')
    os.makedirs(outquery+'Output Citra Query'+'\\')




pathexcel = 'DataQueryExcel.xlsx'
if not os.path.exists(pathexcel):
    wb = Workbook()
    ws = wb.active
    ws.title = str(preprocess)
    wb.save(pathexcel)
wb = openpyxl.load_workbook(pathexcel)


if not str(preprocess) in wb.sheetnames:
    print('waga')
    wb.create_sheet(str(preprocess))
    wb.save(pathexcel)



sheet = wb[str(preprocess)]


print("AFTER"+str(wb.sheetnames))


print("\n\n=======Running=======\n")
print(str(preprocess) + "\n")
print(str(mode)+"\n")
print(str(cnn) + "\n")
print("=======================")
with open(outquery + 'Output Akurasi Query\\Akurasi ' + str(mode) + ' ' + str(maxres) + '.txt', 'w') as g:
    for n in gambar[0:]:
        gc.collect()
        urutan = urutan + 1
        namakd = alamatinput + (gambar[urutan]).split(".")[0] + '_Scale_100%_Rotasi_30.jpg'
        cekkd = (gambar[urutan]).split(".")[0]
        print(namakd)

        queryVec = ekstrak(namakd, cnn)

        cek = []

        ##========BERHASIL==========
        if mode == 'Manhattan':
            dists = np.linalg.norm(feats - queryVec, ord=1,
                                   axis=1)  # L1 distances to features (Manhattan) @89.286  @76.789 @67.857
        elif mode == 'Euclidean':
            dists = np.linalg.norm(feats - queryVec, ord=2,
                                   axis=1)  # L2 distances to features (Euclidean) @71.429 @76.786 @73.2143
        elif mode == 'Chebyshev':
            dists = np.linalg.norm(feats - queryVec, ord=np.inf,
                                   axis=1)  # Max distances to features (Chebesvy) @41.0714 @60.714 @55.357
        elif mode == 'Minkowski':
            dists = np.linalg.norm(feats - queryVec, ord=3,
                                   axis=1)  #Minkowski with Order = 3
        elif mode == 'Vector Cosine':
            dists = 1 - (np.dot(feats, queryVec) / (
                        np.linalg.norm(feats) * np.linalg.norm(queryVec)))  # Cosine Similarity @71.429 @76.786 @73.2143

        # dists = np.sum(np.abs(feats-queryVec)**2,axis=-1)**(1./2) #Euclidean Distance versi alternatif

        # dists = np.linalg.norm(feats-queryVec,ord=-1, axis=1)  # L2 distances to features

        ##==========================
        ids = np.argsort(dists)[:maxres]  # Top 30 results
        scores = [(dists[id], imgNames[id]) for id in ids]

        # print(scores)

        rank_ID = ids

        datab = {}  # database semua


        cekkd = namakd.split("\\")[1]
        cekkd2 = cekkd.split(".")[0]

        cekkd = cekkd.split("_")[0]

        datab["Rata2_" + str(cekkd)] = 0

        imlist = []

        with open(outquery + '\\Output Text Query\\' + cekkd2 + '.txt', 'w') as f:
            for i, index in enumerate(rank_ID[0:maxres]):
                imlist.append(imgNames[index])
                #   print("image names: "+str(imgNames[index]) + " scores: "+str(scores[i]))
                print("image names: " + str(imgNames[index].decode()))
                f.write(str(imgNames[index].decode()) + '\n')
                # ===========PROSES MENGHITUNG==========

                namakd3 = str(imgNames[index].decode())
                namakd3 = (namakd3.split("_"))[0]
                print(cekkd, " = ", (namakd3.split("_"))[0])
                if (cekkd == (namakd3.split("_"))[0]):
                    datab["Rata2_" + str(cekkd)] = int(datab.get(str("Rata2_" + str(cekkd)))) + 1
            print(datab)
        datab["Rata2_" + str(cekkd)] = (float(datab.get(str("Rata2_" + str(cekkd)))) / maxres) * 100
        print(datab)
        plt.rcParams.update({'font.size': 5})
        fig = plt.figure(figsize=(8, 8))
        # print("top %d images in order are: " %maxres,str( imlist,'utf-8'))
        if ((float(datab.get(str("Rata2_" + str(cekkd))))) >= cap):
            hit = hit + 1
            hit2 = hit2 + datab.get(str("Rata2_" + str(cekkd)))
            g.write(str(cekkd2) + ' = ' + str(datab.get("Rata2_" + str(cekkd))) + '\n')
            if (pil%2==0):
                sheet.cell(row=1, column=((4*modelkeras)-2)).value = str(mode)
                sheet.cell(row=1, column=(((4 * modelkeras) - 2)+1)).value = (str(cnn))
                sheet.cell(row=hit, column=((4 * modelkeras) - 2)).value = (str(cekkd2.split("_")[0]) + " " + str(cekkd2.split("_")[1]))
                sheet.cell(row=hit, column=(((4 * modelkeras) - 2)+1)).value = (str(datab.get("Rata2_" + str(cekkd))))
            else:
                sheet.cell(row=1, column=((4 * modelkeras))).value = str(mode)
                sheet.cell(row=1, column=(((4 * modelkeras) + 1))).value = (str(cnn))
                sheet.cell(row=hit, column=((4 * modelkeras))).value = (str(cekkd2.split("_")[0]) + " " + str(cekkd2.split("_")[1]))
                sheet.cell(row=hit, column=(((4 * modelkeras) + 1))).value = (str(datab.get("Rata2_" + str(cekkd))))


            queryImg = mpimg.imread(namakd)

            plt.subplot(int(rows), int(columns), int(columnsz))
            plt.title(
                "Query Image:\n" + (namakd.split("\\")[1]).split("_")[0] + "\nAccuracy: " + str(
                    datab.get(str("Rata2_" + str(cekkd)))) + "%")
            ax = plt.gca()
            ax.axes.xaxis.set_visible(False)
            ax.axes.yaxis.set_visible(False)

            plt.grid(True)
            plt.imshow(queryImg, cmap='gray')

            j = columns
            for i, im in enumerate(imlist):
                j = j + 1
                img2 = mpimg.imread(path + str(im, 'utf-8'))
                # img2.mean(axis=2)
                fig.add_subplot(rows, columns, j)
                nama = str(imlist[i].decode()).split("_")
                plt.title(nama[0] + "\n" + nama[3] + "\n" + nama[5], y=1)
                # print("ASDASDAS : "+nama[0] + " | R" + "\n" + (nama[3].split(" "))[0] + "\n" + nama[4] + "°\n" + (nama[6].split(".")[0]))
                ax = plt.gca()
                ax.axes.xaxis.set_visible(False)
                ax.axes.yaxis.set_visible(False)

                plt.grid(True)
                plt.imshow(img2, cmap='gray')

            plt.tight_layout()
            plt.savefig(outquery + 'Output Citra Query\\' + cekkd2 + '_' + '.png')
            plt.clf()
            if enable_lock == 1:
                if (float(datab.get(str("Rata2_" + str(cekkd))))) >= cap and (float(datab.get(str("Rata2_" + str(cekkd))))) < 70:  # PENTING! HANYA UNTUK KASUS KHUSUS MATIKAN JIKA BUKAN 56 MINKOWSKI
                    lock = lock + 1  # =============PENTING! HANYA UNTUK KASUS KHUSUS MATIKAN JIKA BUKAN 56 MINKOWSKI==========
                    print("Tambah Lock")
                    if lock == stop:
                        print("OLD CAP == " + str(cap))
                        cap = cap+1  # =============PENTING! HANYA UNTUK KASUS KHUSUS MATIKAN JIKA BUKAN 56 MINKOWSKI============
                        enable_lock = 0
                        print("NEW CAP == " + str(cap))
                        print("==Disable Lock[1]==")
                elif (float(datab.get(str("Rata2_" + str(cekkd))))) >= cap and (float(datab.get(str("Rata2_" + str(cekkd))))) <= topcap:  # PENTING! HANYA UNTUK KASUS KHUSUS MATIKAN JIKA BUKAN 56 MINKOWSKI
                    lock = lock + 1  # =============PENTING! HANYA UNTUK KASUS KHUSUS MATIKAN JIKA BUKAN 56 MINKOWSKI==========
                    print("Tambah Lock")
                    if lock == stop:
                        print("OLD CAP == " + str(cap))
                        cap = cap+1  # =============PENTING! HANYA UNTUK KASUS KHUSUS MATIKAN JIKA BUKAN 56 MINKOWSKI============
                        enable_lock = 0
                        print("NEW CAP == "+ str(cap))
                        print("==Disable Lock[2]==")
        else:
            print("LESS THAN "+str(cap))
            if os.path.exists(outquery + 'Output Text Query\\' + cekkd2 + '.txt'):
                os.remove(outquery + 'Output Text Query\\' + cekkd2 + '.txt')
        print("================URUTAN ", urutan+1, "==================")
    print(" Total Hasil ="+str(hit2)+" || Total Ditemukan="+str(hit))
    print("Mean =" + str(hit2/hit))
    g.write("===============\n")
    #worksheet.write( hit+1,pil * 2, "===============")
    if(hit == 0):
        g.write("GAMBAR TIDAK DITEMUKAN DIATAS "+str(cap)+"%"+ "\n")
        if (pil % 2 == 0):
            sheet.cell(hit + 1, ((4 * modelkeras) - 2)).value = "TOTAL"
            sheet.cell(hit + 1, ((4 * modelkeras) - 2) + 1).value = hit
            sheet.cell(hit + 2, ((4 * modelkeras) - 2)).value = "MEAN"
            sheet.cell(hit + 2,((4*modelkeras)-2) + 1).value = "GAMBAR TIDAK DITEMUKAN DIATAS "+str(cap)+"%"
        else:
            sheet.cell(hit + 1, ((4 * modelkeras))).value = "TOTAL"
            sheet.cell(hit + 1, ((4 * modelkeras)) + 1).value = hit
            sheet.cell(hit + 2, ((4 * modelkeras))).value = "MEAN"
            sheet.cell(hit + 2,((4*modelkeras) + 1)).value = "GAMBAR TIDAK DITEMUKAN DIATAS "+str(cap)+"%"
    else:
        g.write(str(hit2 / hit) + "\n")
        if (pil % 2 == 0):
            sheet.cell(hit + 1, ((4 * modelkeras) - 2)).value = "TOTAL"
            sheet.cell(hit + 1, ((4 * modelkeras) - 2) + 1).value = hit
            sheet.cell(hit + 2, ((4 * modelkeras) - 2)).value = "MEAN"
            sheet.cell(hit + 2, (((4 * modelkeras)-2) + 1)).value = hit2/hit

            sheet.cell(59, 1).value = str(mode)
            sheet.cell(60, 1).value = "Model"
            sheet.cell(60, 1+1).value = "Mean"
            sheet.cell(60, 1+2).value = "Total"
            sheet.cell(60 + modelkeras, 1).value = str(cnn)
            sheet.cell(60+modelkeras, 1+1).value = hit2/hit
            sheet.cell(60+modelkeras, 1+2).value = hit

        else:
            sheet.cell(hit + 1, ((4 * modelkeras))).value = "TOTAL"
            sheet.cell(hit + 1, ((4 * modelkeras)) + 1).value = hit
            sheet.cell(hit + 2, ((4 * modelkeras))).value = "MEAN"
            sheet.cell(hit + 2, ((4 * modelkeras) + 1)).value = hit2/hit

            sheet.cell(69, 1).value = str(mode)
            sheet.cell(70, 1).value = "Model"
            sheet.cell(70, 1 + 1).value = "Mean"
            sheet.cell(70, 1 + 2).value = "Total"
            sheet.cell(70 + modelkeras, 1).value = str(cnn)
            sheet.cell(70 + modelkeras, 1 + 1).value = hit2 / hit
            sheet.cell(70 + modelkeras, 1 + 2).value = hit

    g.write("===============\n")
    #worksheet.write( hit + 3,pil * 2, "===============")
    g.close()
    wb.title = str(preprocess)
    wb.save(pathexcel)
    print("===Finished Running:===\n")
    print(str(preprocess) + "\n")
    print(str(mode)+"\n")
    print(str(cnn) + "\n")
    print("=======================")








